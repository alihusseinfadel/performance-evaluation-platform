import streamlit as st
import datetime
import os
import io
import json
import base64
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go

st.set_page_config(
    page_title="منصة تقييم الأداء الذكي - جامعة ديالى",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(APP_DIR, "data")

# ─────────────────────────────────────────────
# Session state defaults
# ─────────────────────────────────────────────
if "dark_mode" not in st.session_state:
    st.session_state.dark_mode = False
if "admin_authenticated" not in st.session_state:
    st.session_state.admin_authenticated = False

# ─────────────────────────────────────────────
# Grade mapping
# ─────────────────────────────────────────────
GRADE_MAP = [
    (90, 100, "ممتاز", "#16a34a"),
    (80, 89, "جيد جداً", "#2563eb"),
    (70, 79, "جيد", "#7c3aed"),
    (60, 69, "مقبول", "#d97706"),
    (0, 59, "ضعيف", "#dc2626"),
]


def score_to_grade(score):
    if score is None:
        return "—", "#888"
    for lo, hi, label, color in GRADE_MAP:
        if lo <= score <= hi:
            return label, color
    return "—", "#888"


# ─────────────────────────────────────────────
# AI Engine
# ─────────────────────────────────────────────
def ai_trend_analysis(scores_by_year):
    """Analyze performance trend from a list of (year_name, score) tuples sorted chronologically."""
    if len(scores_by_year) < 2:
        return {"trend": "غير كافٍ", "icon": "➖", "color": "#888",
                "description": "تحتاج سنتين على الأقل لتحليل الاتجاه"}
    values = [s for _, s in scores_by_year]
    diffs = [values[i+1] - values[i] for i in range(len(values)-1)]
    avg_diff = sum(diffs) / len(diffs)
    last_diff = diffs[-1]

    if avg_diff > 5 and last_diff > 0:
        return {"trend": "تحسن ملحوظ", "icon": "🚀", "color": "#16a34a",
                "description": f"تحسن بمعدل {avg_diff:+.1f} نقطة سنوياً"}
    elif avg_diff > 0:
        return {"trend": "تحسن تدريجي", "icon": "📈", "color": "#22c55e",
                "description": f"تحسن طفيف بمعدل {avg_diff:+.1f} نقطة سنوياً"}
    elif avg_diff > -3:
        return {"trend": "مستقر", "icon": "➡️", "color": "#2563eb",
                "description": "أداء مستقر مع تغيرات طفيفة"}
    elif avg_diff > -8:
        return {"trend": "تراجع طفيف", "icon": "📉", "color": "#d97706",
                "description": f"تراجع بمعدل {avg_diff:+.1f} نقطة سنوياً"}
    else:
        return {"trend": "تراجع حاد", "icon": "⚠️", "color": "#dc2626",
                "description": f"تراجع كبير بمعدل {avg_diff:+.1f} نقطة سنوياً — يحتاج متابعة عاجلة"}


def ai_predict_next(scores_by_year):
    """Simple linear regression prediction for next year."""
    if len(scores_by_year) < 2:
        return None
    values = [s for _, s in scores_by_year]
    n = len(values)
    x = list(range(n))
    x_mean = sum(x) / n
    y_mean = sum(values) / n
    num = sum((x[i] - x_mean) * (values[i] - y_mean) for i in range(n))
    den = sum((x[i] - x_mean) ** 2 for i in range(n))
    if den == 0:
        return round(y_mean)
    slope = num / den
    intercept = y_mean - slope * x_mean
    predicted = slope * n + intercept
    return max(0, min(100, round(predicted)))


def ai_self_gap_alert(self_score, sup_score):
    """Detect gap between self and supervisor evaluation."""
    if self_score is None or sup_score is None:
        return None
    gap = self_score - sup_score
    if gap > 15:
        return {"type": "overconfidence", "icon": "🔴", "color": "#dc2626",
                "msg": f"تقييم ذاتي أعلى بـ {gap} نقطة — مبالغة في التقييم الذاتي"}
    elif gap > 8:
        return {"type": "slight_overconfidence", "icon": "🟡", "color": "#d97706",
                "msg": f"تقييم ذاتي أعلى بـ {gap} نقطة — فرق ملحوظ"}
    elif gap < -15:
        return {"type": "underconfidence", "icon": "🔵", "color": "#2563eb",
                "msg": f"تقييم ذاتي أقل بـ {abs(gap)} نقطة — تواضع مفرط"}
    elif gap < -8:
        return {"type": "slight_underconfidence", "icon": "🔵", "color": "#7c3aed",
                "msg": f"تقييم ذاتي أقل بـ {abs(gap)} نقطة — يقلل من أدائه"}
    return {"type": "aligned", "icon": "✅", "color": "#16a34a",
            "msg": f"فرق {abs(gap)} نقاط فقط — تقييم متوازن"}


def ai_smart_rank(associates, evaluations, years):
    """Rank employees considering consistency, improvement, and average."""
    rankings = []
    for a in associates:
        person_evals = sorted(
            [e for e in evaluations if e.get("associate_id") == a["id"] and e.get("supervisor_score") is not None],
            key=lambda e: e.get("year_id", 0)
        )
        if not person_evals:
            continue
        scores = [e["supervisor_score"] for e in person_evals]
        avg = sum(scores) / len(scores)

        # Consistency (lower std dev = more consistent)
        if len(scores) > 1:
            mean = avg
            variance = sum((s - mean) ** 2 for s in scores) / len(scores)
            std_dev = variance ** 0.5
            consistency = max(0, 100 - std_dev * 3)
        else:
            consistency = 50

        # Improvement rate
        if len(scores) >= 2:
            improvement = scores[-1] - scores[0]
        else:
            improvement = 0

        # Weighted composite: 50% avg, 25% consistency, 25% improvement
        composite = avg * 0.50 + consistency * 0.25 + min(max(improvement + 50, 0), 100) * 0.25

        # Category
        if avg >= 90 and improvement >= 0:
            category = "🌟 نجم متميز"
            cat_color = "#16a34a"
        elif improvement > 5:
            category = "🚀 نجم صاعد"
            cat_color = "#22c55e"
        elif avg >= 75 and abs(improvement) <= 5:
            category = "✅ أداء مستقر"
            cat_color = "#2563eb"
        elif improvement < -5:
            category = "⚠️ يحتاج متابعة"
            cat_color = "#d97706"
        elif avg < 60:
            category = "🔴 أداء ضعيف"
            cat_color = "#dc2626"
        else:
            category = "➡️ أداء عادي"
            cat_color = "#888"

        rankings.append({
            "id": a["id"],
            "name": a["name"],
            "avg": round(avg, 1),
            "consistency": round(consistency, 1),
            "improvement": round(improvement, 1),
            "composite": round(composite, 1),
            "category": category,
            "cat_color": cat_color,
            "eval_count": len(scores),
        })

    rankings.sort(key=lambda r: r["composite"], reverse=True)
    return rankings


def ai_recommendations(assoc_name, avg_score, trend, gap_alert, category):
    """Generate smart recommendations per employee."""
    recs = []
    if avg_score >= 90:
        recs.append("🏆 مرشح للتكريم والمكافأة السنوية")
        recs.append("📋 يُنصح بإسناد مهام قيادية إضافية")
    elif avg_score >= 80:
        recs.append("📈 أداء جيد جداً — يُنصح بدعمه للوصول للتميز")
    elif avg_score >= 70:
        recs.append("📝 يُنصح بحضور دورات تطويرية لتحسين الأداء")
    elif avg_score >= 60:
        recs.append("⚡ يحتاج خطة تحسين أداء فورية")
        recs.append("👨‍🏫 يُنصح بمتابعة شهرية من المشرف المباشر")
    else:
        recs.append("🚨 أداء ضعيف — يحتاج تدخل إداري عاجل")
        recs.append("📋 يُنصح بوضع خطة إنذار مبكر")

    if trend and trend.get("trend") == "تراجع حاد":
        recs.append("📉 اتجاه تراجعي — يُنصح بجلسة تقييم خاصة")
    elif trend and trend.get("trend") == "تحسن ملحوظ":
        recs.append("🌟 في مسار تصاعدي — يُنصح بالتشجيع والتحفيز")

    if gap_alert and gap_alert.get("type") == "overconfidence":
        recs.append("🔍 فجوة كبيرة في التقييم الذاتي — يُنصح بجلسة توعية")
    elif gap_alert and gap_alert.get("type") == "underconfidence":
        recs.append("💡 يقلل من أدائه — يُنصح بتعزيز الثقة والتقدير")

    return recs


# ─────────────────────────────────────────────
# PDF Report Generator
# ─────────────────────────────────────────────
def generate_employee_pdf(assoc, evaluations, years, affiliations, certificates, nicknames, positions):
    """Generate a professional PDF evaluation report for an employee."""
    try:
        from fpdf import FPDF
        import arabic_reshaper
        from bidi.algorithm import get_display

        def ar(text):
            try:
                reshaped = arabic_reshaper.reshape(str(text))
                return get_display(reshaped)
            except Exception:
                return str(text)

        class PDF(FPDF):
            def header(self):
                self.set_font("tahoma", "B", 14)
                self.cell(0, 10, ar("منصة تقييم الأداء الذكي — جامعة ديالى"), 0, 1, "C")
                self.set_font("tahoma", "", 10)
                self.cell(0, 6, ar(f"تقرير تقييم الأداء — {datetime.date.today()}"), 0, 1, "C")
                self.line(10, self.get_y() + 2, 200, self.get_y() + 2)
                self.ln(6)

            def footer(self):
                self.set_y(-15)
                self.set_font("tahoma", "", 8)
                self.cell(0, 10, ar(f"صفحة {self.page_no()}"), 0, 0, "C")

        pdf = PDF()
        font_path = os.path.join(APP_DIR, "tahoma.ttf")
        if os.path.exists(font_path):
            pdf.add_font("tahoma", "", font_path, uni=True)
            pdf.add_font("tahoma", "B", font_path, uni=True)
        else:
            pdf.add_font("tahoma", "", fname="", uni=True)

        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=20)

        # Employee info
        pdf.set_font("tahoma", "B", 13)
        pdf.cell(0, 10, ar(f"المنتسب: {assoc['name']}"), 0, 1, "R")
        pdf.set_font("tahoma", "", 11)

        aff = lookup_name(affiliations, assoc.get("affiliation_id"))
        cert = lookup_name(certificates, assoc.get("certificate_id"))
        nick = lookup_name(nicknames, assoc.get("nickname_id"))
        pos = lookup_name(positions, assoc.get("position_id"))

        info_lines = [
            f"الجهة: {aff}",
            f"المنصب: {pos}",
            f"الشهادة: {cert}",
            f"اللقب العلمي: {nick}",
            f"الهاتف: {assoc.get('phone', '—')}",
        ]
        for line in info_lines:
            pdf.cell(0, 8, ar(line), 0, 1, "R")

        pdf.ln(5)

        # Evaluations table
        person_evals = [e for e in evaluations if e.get("associate_id") == assoc["id"] and e.get("supervisor_score") is not None]

        if person_evals:
            pdf.set_font("tahoma", "B", 12)
            pdf.cell(0, 10, ar("سجل التقييمات"), 0, 1, "R")

            # Table header
            pdf.set_fill_color(30, 64, 175)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("tahoma", "B", 10)
            col_w = [50, 35, 35, 35, 35]
            headers = ["السنة", "التقييم الذاتي", "تقييم المشرف", "التقدير", "الحالة"]
            for i, h in enumerate(headers):
                pdf.cell(col_w[i], 9, ar(h), 1, 0, "C", True)
            pdf.ln()

            pdf.set_text_color(0, 0, 0)
            pdf.set_font("tahoma", "", 10)
            for e in person_evals:
                yr_name = lookup_name(years, e["year_id"], "year")
                g, _ = score_to_grade(e["supervisor_score"])
                self_s = str(e.get("self_score", "—")) if e.get("self_score") is not None else "—"
                pdf.cell(col_w[0], 8, ar(yr_name), 1, 0, "C")
                pdf.cell(col_w[1], 8, ar(self_s), 1, 0, "C")
                pdf.cell(col_w[2], 8, str(e["supervisor_score"]), 1, 0, "C")
                pdf.cell(col_w[3], 8, ar(g), 1, 0, "C")
                pdf.cell(col_w[4], 8, ar(e.get("status", "—")), 1, 0, "C")
                pdf.ln()

            # Cumulative
            avg = round(sum(e["supervisor_score"] for e in person_evals) / len(person_evals), 1)
            cum_grade, _ = score_to_grade(int(avg))
            pdf.ln(5)
            pdf.set_font("tahoma", "B", 12)
            pdf.cell(0, 10, ar(f"المعدل التراكمي: {avg} ({cum_grade})"), 0, 1, "R")

        # Signature area
        pdf.ln(20)
        pdf.set_font("tahoma", "", 10)
        pdf.cell(90, 8, ar("توقيع المشرف: _______________"), 0, 0, "C")
        pdf.cell(90, 8, ar("التاريخ: _______________"), 0, 1, "C")

        buf = io.BytesIO()
        pdf.output(buf)
        return buf.getvalue()
    except Exception as e:
        st.error(f"خطأ في توليد PDF: {e}")
        return None


# ─────────────────────────────────────────────
# Global search helper
# ─────────────────────────────────────────────
def search_associates(associates, query, affiliations, certificates, nicknames, positions):
    """Search associates by name, phone, or any related field."""
    if not query:
        return associates
    q = query.strip().lower()
    results = []
    for a in associates:
        searchable = " ".join([
            a.get("name", ""),
            a.get("phone", ""),
            lookup_name(affiliations, a.get("affiliation_id")),
            lookup_name(certificates, a.get("certificate_id")),
            lookup_name(nicknames, a.get("nickname_id")),
            lookup_name(positions, a.get("position_id")),
        ]).lower()
        if q in searchable:
            results.append(a)
    return results


# ─────────────────────────────────────────────
# Progress bar helper
# ─────────────────────────────────────────────
def render_progress_bar(value, max_val, label="", color="#2563eb"):
    pct = round(value / max_val * 100) if max_val > 0 else 0
    return f"""
    <div style="margin: 0.5rem 0;">
        <div style="display:flex; justify-content:space-between; margin-bottom:4px;">
            <span style="font-size:0.9rem; color:#64748b;">{label}</span>
            <span style="font-size:0.9rem; font-weight:700; color:{color};">{pct}%</span>
        </div>
        <div style="background:#e2e8f0; border-radius:10px; height:10px; overflow:hidden;">
            <div style="background:linear-gradient(90deg, {color}, {color}cc); width:{pct}%; height:100%;
                        border-radius:10px; transition: width 1s ease;"></div>
        </div>
    </div>
    """


# ─────────────────────────────────────────────
# Employee profile card
# ─────────────────────────────────────────────
def render_profile_card(assoc, affiliations, certificates, nicknames, positions, evaluations, years):
    aff = lookup_name(affiliations, assoc.get("affiliation_id"))
    cert = lookup_name(certificates, assoc.get("certificate_id"))
    nick = lookup_name(nicknames, assoc.get("nickname_id"))
    pos = lookup_name(positions, assoc.get("position_id"))

    person_evals = [e for e in evaluations if e.get("associate_id") == assoc["id"] and e.get("supervisor_score") is not None]
    avg = round(sum(e["supervisor_score"] for e in person_evals) / len(person_evals), 1) if person_evals else 0
    g, gc = score_to_grade(int(avg)) if avg else ("—", "#888")
    eval_count = len(person_evals)
    initials = assoc["name"][:2] if assoc.get("name") else "؟"

    return f"""
    <div class="result-card" style="border-right-color: {gc};">
        <div style="display:flex; align-items:center; gap:15px; flex-direction:row-reverse;">
            <div style="width:55px;height:55px;border-radius:50%;
                        background:linear-gradient(135deg,{gc},{gc}88);
                        display:flex;align-items:center;justify-content:center;
                        color:#fff;font-size:1.2rem;font-weight:800;flex-shrink:0;">
                {initials}
            </div>
            <div style="flex:1;">
                <h4 style="margin:0 0 4px 0;">{assoc['name']}</h4>
                <p style="margin:0; font-size:0.9rem !important; color:#64748b !important;">
                    {nick} | {pos} | {aff} | {cert}
                </p>
                <p style="margin:4px 0 0 0;">
                    المعدل: <span class="eval-badge" style="background:{gc}18;color:{gc};">{avg} - {g}</span>
                    &nbsp; | &nbsp; عدد التقييمات: <strong>{eval_count}</strong>
                    &nbsp; | &nbsp; 📞 {assoc.get('phone', '—')}
                </p>
            </div>
        </div>
    </div>
    """


# ─────────────────────────────────────────────
# JSON persistence helpers
# ─────────────────────────────────────────────
def _path(name):
    return os.path.join(DATA_DIR, name)


def load_json(name):
    p = _path(name)
    if os.path.exists(p):
        with open(p, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def save_json(name, data):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(_path(name), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def next_id(items):
    return max((it.get("id", 0) for it in items), default=0) + 1


# ─────────────────────────────────────────────
# Data loaders (cached in session)
# ─────────────────────────────────────────────
def load_all():
    d = {}
    d["affiliations"] = load_json("affiliations.json")
    d["certificates"] = load_json("certificates.json")
    d["nicknames"] = load_json("nicknames.json")
    d["positions"] = load_json("positions.json")
    d["associates"] = load_json("associates.json")
    d["years"] = load_json("evaluation_years.json")
    d["evaluations"] = load_json("evaluations.json")
    return d


def lookup_name(items, item_id, key="name"):
    for it in items:
        if it.get("id") == item_id:
            return it.get(key, "—")
    return "—"


# ─────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────
def export_df_excel(df, sheet_name="Sheet1"):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.sheet_view.rightToLeft = True

        h_fill = PatternFill("solid", fgColor="1E40AF")
        h_font = Font(name="Tahoma", bold=True, color="FFFFFF", size=10)
        h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        cols = list(df.columns)
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.fill = h_fill; c.font = h_font; c.alignment = h_align; c.border = border

        d_font = Font(name="Tahoma", size=10)
        d_align = Alignment(horizontal="right", vertical="center")
        fills = [PatternFill("solid", fgColor="EFF6FF"), PatternFill("solid", fgColor="F8FAFF")]

        for ri, (_, row) in enumerate(df.iterrows(), 2):
            for ci, col in enumerate(cols, 1):
                c = ws.cell(row=ri, column=ci, value=row[col])
                c.fill = fills[ri % 2]; c.font = d_font; c.alignment = d_align; c.border = border

        for ci, col in enumerate(cols, 1):
            max_len = max(len(str(col)), df[col].astype(str).str.len().max() if len(df) > 0 else 0)
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)

        ws.row_dimensions[1].height = 35
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception as e:
        st.error(f"خطأ في التصدير: {e}")
        return None


# ─────────────────────────────────────────────
# Logo / Watermark
# ─────────────────────────────────────────────
@st.cache_data
def get_logo_b64():
    try:
        p = os.path.join(APP_DIR, "diyala_full_seal.png")
        if os.path.exists(p):
            with open(p, "rb") as f:
                return base64.b64encode(f.read()).decode("ascii")
    except Exception:
        pass
    return None


LOGO_B64 = get_logo_b64()

# ─────────────────────────────────────────────
# CSS — identical to A3
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@300;400;500;700;800&display=swap');

/* ── Base ── */
* { font-family: 'Tajawal', sans-serif !important; }
html, body, [data-testid="stAppViewContainer"] { font-size: 17px !important; }
.main .block-container { padding-top: 0.5rem; max-width: 1200px; }
.main { background: linear-gradient(135deg, #f0f4f8 0%, #e8edf4 100%) !important; }
h1,h2,h3,h4,h5,h6,p,li,label,span,div { direction: rtl; text-align: right; }
h1 { font-size: 2rem !important; }
h2 { font-size: 1.7rem !important; }
h3 { font-size: 1.45rem !important; }
h4 { font-size: 1.25rem !important; }
p, li, label, span { font-size: 1.05rem !important; }
input, select, textarea, button { font-size: 1.05rem !important; }
[data-testid="stMarkdownContainer"] p { font-size: 1.05rem !important; line-height: 1.9; }
[data-testid="stDataFrame"] { font-size: 0.95rem !important; }
.stTabs [data-baseweb="tab"] { font-size: 1.05rem !important; padding: 0.7rem 1.2rem !important; }

/* ── Animations ── */
@keyframes fadeInUp {
    from { opacity: 0; transform: translateY(20px); }
    to { opacity: 1; transform: translateY(0); }
}
@keyframes slideInRight {
    from { opacity: 0; transform: translateX(30px); }
    to { opacity: 1; transform: translateX(0); }
}
@keyframes pulse {
    0%, 100% { transform: scale(1); }
    50% { transform: scale(1.02); }
}
@keyframes shimmer {
    0% { background-position: -200% 0; }
    100% { background-position: 200% 0; }
}

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    direction: rtl; text-align: right;
    background: linear-gradient(180deg, #0f2744 0%, #091a2e 50%, #060e1a 100%) !important;
    min-width: 310px !important; width: 310px !important;
    border-left: 1px solid rgba(59,130,246,0.15);
}
[data-testid="stSidebar"] > div:first-child { min-width: 310px !important; width: 310px !important; }
[data-testid="stSidebar"] *,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] div { color: #c8ddf0 !important; }
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 { color: #fff !important; }

[data-testid="stSidebar"] .stRadio > div > label {
    background: rgba(255,255,255,0.03); color: #c0d8ef !important;
    border-radius: 10px; padding: 0.75rem 1.1rem !important;
    transition: all 0.3s cubic-bezier(0.4,0,0.2,1);
    border: 1px solid rgba(255,255,255,0.05);
    font-size: 1.05rem !important;
    backdrop-filter: blur(4px);
}
[data-testid="stSidebar"] .stRadio > div > label:hover {
    background: rgba(59,130,246,0.12);
    border-color: rgba(59,130,246,0.25);
    transform: translateX(-3px);
}
[data-testid="stSidebar"] .stRadio label:has(input:checked) {
    background: linear-gradient(135deg, #2563eb 0%, #1d4ed8 100%) !important;
    color: #fff !important;
    box-shadow: 0 4px 20px rgba(37,99,235,0.35), inset 0 1px 0 rgba(255,255,255,0.15);
    border-color: transparent;
}
[data-testid="stSidebar"] .stRadio label:has(input:checked) p { color: #fff !important; }
[data-testid="stSidebar"] .stRadio > div > label > div:first-child { display: none; }

/* ── Hide defaults ── */
[data-testid="InputInstructions"] { display: none !important; }
#MainMenu, footer, header[data-testid="stHeader"] { visibility: hidden; }
[data-testid="stToolbar"] { display: none !important; }

/* ── Result Card — animated ── */
.result-card {
    background: #fff; border: 1px solid #e2e8f0; border-radius: 14px;
    padding: 1.3rem 1.5rem; margin: 0.6rem 0;
    border-right: 5px solid #2563eb;
    box-shadow: 0 2px 12px rgba(0,0,0,0.04), 0 1px 3px rgba(0,0,0,0.06);
    transition: all 0.3s cubic-bezier(0.4,0,0.2,1);
    animation: fadeInUp 0.4s ease-out;
}
.result-card:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 25px rgba(37,99,235,0.1), 0 3px 10px rgba(0,0,0,0.06);
    border-right-color: #1d4ed8;
}
.result-card h4 { color: #1e40af; margin: 0 0 0.5rem 0; font-size: 1.2rem !important; font-weight: 700; }
.result-card p  { color: #475569; margin: 0.3rem 0; line-height: 1.9; font-size: 1.05rem !important; }

/* ── Stat Card — glass morphism ── */
.stat-card {
    background: linear-gradient(145deg, #ffffff 0%, #f8fafc 100%);
    border: 1px solid #e2e8f0; border-radius: 16px;
    padding: 1.2rem; text-align: center;
    border-top: 4px solid #2563eb;
    box-shadow: 0 4px 15px rgba(0,0,0,0.04), 0 1px 3px rgba(0,0,0,0.06);
    transition: all 0.3s cubic-bezier(0.4,0,0.2,1);
    animation: fadeInUp 0.5s ease-out;
    position: relative; overflow: hidden;
}
.stat-card::before {
    content: ''; position: absolute; top: 0; left: 0; right: 0; height: 4px;
    background: linear-gradient(90deg, #2563eb, #7c3aed, #2563eb);
    background-size: 200% 100%; animation: shimmer 3s infinite;
}
.stat-card:hover {
    transform: translateY(-4px);
    box-shadow: 0 12px 30px rgba(37,99,235,0.12), 0 4px 10px rgba(0,0,0,0.06);
}
.stat-card h2 { text-align: center; color: #1e40af; margin: 0; font-size: 2.4rem; font-weight: 800; }
.stat-card p  { text-align: center; color: #64748b; margin: 0.3rem 0 0 0; font-size: 1rem; font-weight: 500; }

/* ── Page Title — gradient ── */
.page-title {
    font-size: 1.8rem; font-weight: 800;
    background: linear-gradient(135deg, #1e40af, #2563eb);
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    background-clip: text;
    margin-bottom: 1.2rem;
    border-bottom: 3px solid transparent;
    border-image: linear-gradient(90deg, #2563eb, #7c3aed) 1;
    padding-bottom: 0.6rem; display: inline-block;
    animation: fadeInUp 0.3s ease-out;
}

/* ── Sidebar Hero ── */
.sb-hero {
    background: linear-gradient(135deg, rgba(37,99,235,0.12), rgba(29,78,216,0.06));
    border: 1px solid rgba(59,130,246,0.15);
    border-radius: 16px; padding: 1.4rem; text-align: center; margin: 0.2rem 0 1.2rem 0;
    backdrop-filter: blur(10px);
}
.sb-hero-logo {
    width: 60px; height: 60px;
    background: linear-gradient(135deg, #3b82f6, #1d4ed8, #7c3aed);
    border-radius: 16px; display: flex; align-items: center;
    justify-content: center; color: #fff; font-size: 1.6rem;
    font-weight: 800; margin: 0 auto 0.6rem;
    box-shadow: 0 6px 20px rgba(37,99,235,0.4);
    animation: pulse 3s infinite;
}
.sb-hero h1 { color: #fff !important; margin: 0 !important; font-size: 1.15rem !important; text-align: center !important; font-weight: 700 !important; }
.sb-hero p  { color: #93c5fd !important; margin: 4px 0 0 0 !important; font-size: 0.85rem !important; text-align: center !important; }

.sb-label {
    color: #93c5fd !important; font-size: 0.78rem; font-weight: 800;
    letter-spacing: 1.2px; margin: 1.4rem 0 0.6rem 0.3rem;
    display: flex; align-items: center; gap: 8px; flex-direction: row-reverse;
    text-transform: uppercase;
}
.sb-label-bar { width: 3px; height: 16px; background: linear-gradient(180deg, #3b82f6, #7c3aed); border-radius: 2px; display: inline-block; }

.sb-footer { margin-top: 1.5rem; padding: 1rem; text-align: center; border-top: 1px solid rgba(255,255,255,0.06); }
.sb-footer p { color: #4a6a8a !important; font-size: 0.82rem !important; margin: 2px 0 !important; text-align: center !important; }

.sb-credits {
    margin-top: 1rem; padding: 1rem;
    background: linear-gradient(135deg, rgba(37,99,235,0.08), rgba(124,58,237,0.05));
    border: 1px solid rgba(59,130,246,0.12);
    border-radius: 12px; text-align: center;
}
.sb-credits-title { color: #fff !important; font-size: 0.95rem !important; font-weight: 700 !important; line-height: 1.6 !important; }
.sb-credits-sub   { color: #93c5fd !important; font-size: 0.85rem !important; }

/* ── Guide Box — gradient + icon ── */
.guide-box {
    background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 50%, #e0e7ff 100%);
    border: 1px solid #bfdbfe; border-radius: 16px;
    padding: 1.4rem 1.6rem; margin-bottom: 1.2rem;
    box-shadow: 0 2px 10px rgba(37,99,235,0.06);
    animation: fadeInUp 0.4s ease-out;
}
.guide-box h4 { color: #1e40af; margin: 0 0 0.6rem 0; font-size: 1.2rem !important; font-weight: 700; }
.guide-box p  { color: #1e3a6e; margin: 0.3rem 0; font-size: 1.05rem; }

/* ── Eval Badge — pill style ── */
.eval-badge {
    display: inline-block; padding: 5px 16px; border-radius: 24px;
    font-weight: 700; font-size: 1rem; text-align: center;
    letter-spacing: 0.3px;
    box-shadow: 0 2px 6px rgba(0,0,0,0.08);
    transition: all 0.2s;
}
.eval-badge:hover { transform: scale(1.05); }

/* ── KPI Row ── */
.kpi-row { display: flex; gap: 1rem; margin-bottom: 1rem; flex-wrap: wrap; }
.kpi-item {
    flex: 1; min-width: 140px; background: #fff; border-radius: 12px;
    padding: 0.8rem; text-align: center; border: 1px solid #e2e8f0;
    box-shadow: 0 2px 8px rgba(0,0,0,0.03);
}
.kpi-value { font-size: 1.6rem; font-weight: 800; margin: 0; }
.kpi-label { font-size: 0.85rem; color: #64748b; margin: 0; }

/* ── Section Divider ── */
.section-divider {
    display: flex; align-items: center; gap: 12px; margin: 1.5rem 0;
    flex-direction: row-reverse;
}
.section-divider h3 { margin: 0; white-space: nowrap; color: #1e40af; font-weight: 700; }
.section-divider::before {
    content: ''; flex: 1; height: 2px;
    background: linear-gradient(90deg, transparent, #cbd5e1, transparent);
}

/* ── Professional Table Override ── */
[data-testid="stDataFrame"] > div { border-radius: 12px; overflow: hidden; }

/* ── Form Inputs ── */
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input,
[data-testid="stSelectbox"] > div > div {
    border-radius: 10px !important;
    border-color: #cbd5e1 !important;
    transition: all 0.2s;
}
[data-testid="stTextInput"] input:focus,
[data-testid="stNumberInput"] input:focus {
    border-color: #2563eb !important;
    box-shadow: 0 0 0 3px rgba(37,99,235,0.12) !important;
}

/* ── Buttons ── */
.stButton > button {
    border-radius: 10px !important;
    font-weight: 600 !important;
    transition: all 0.3s cubic-bezier(0.4,0,0.2,1) !important;
    letter-spacing: 0.3px;
}
.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 15px rgba(37,99,235,0.2) !important;
}
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #2563eb, #1d4ed8) !important;
}

/* ── Download Buttons ── */
.stDownloadButton > button {
    border-radius: 10px !important;
    font-weight: 600 !important;
    transition: all 0.3s !important;
}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    gap: 4px; border-bottom: 2px solid #e2e8f0;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 10px 10px 0 0 !important;
    font-weight: 600 !important;
}

/* ── Alert Boxes ── */
.alert-card {
    background: linear-gradient(135deg, #fef2f2, #fee2e2);
    border: 1px solid #fecaca; border-radius: 14px;
    padding: 1rem 1.3rem; margin: 0.5rem 0;
    border-right: 5px solid #dc2626;
    animation: slideInRight 0.4s ease-out;
}
.alert-card h4 { color: #991b1b; margin: 0 0 0.3rem 0; }
.alert-card p { color: #7f1d1d; margin: 0; }

/* ── Success Card ── */
.success-card {
    background: linear-gradient(135deg, #f0fdf4, #dcfce7);
    border: 1px solid #bbf7d0; border-radius: 14px;
    padding: 1rem 1.3rem; margin: 0.5rem 0;
    border-right: 5px solid #16a34a;
}

/* ── Print Ready ── */
@media print {
    [data-testid="stSidebar"] { display: none !important; }
    .main .block-container { max-width: 100% !important; padding: 0 !important; }
    .stat-card, .result-card { break-inside: avoid; }
    .stButton, .stDownloadButton { display: none !important; }
}
</style>
""", unsafe_allow_html=True)

# Dark mode CSS injection
if st.session_state.dark_mode:
    st.markdown("""
    <style>
    .main { background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%) !important; }
    .result-card { background: #1e293b; border-color: #334155; }
    .result-card h4 { color: #93c5fd !important; }
    .result-card p { color: #cbd5e1 !important; }
    .stat-card { background: linear-gradient(145deg, #1e293b, #0f172a); border-color: #334155; }
    .stat-card p { color: #94a3b8 !important; }
    .guide-box { background: linear-gradient(135deg, #1e3a5f, #172554) !important; border-color: #1e40af !important; }
    .guide-box h4 { color: #93c5fd !important; }
    .guide-box p { color: #bfdbfe !important; }
    h1,h2,h3,h4 { color: #e2e8f0 !important; }
    p, li, label, span { color: #cbd5e1 !important; }
    [data-testid="stDataFrame"] { color: #e2e8f0 !important; }
    .alert-card { background: linear-gradient(135deg, #450a0a, #7f1d1d) !important; border-color: #991b1b !important; }
    .alert-card h4, .alert-card p { color: #fecaca !important; }
    .success-card { background: linear-gradient(135deg, #052e16, #166534) !important; border-color: #16a34a !important; }
    </style>
    """, unsafe_allow_html=True)

if LOGO_B64:
    st.markdown(
        '<style>'
        '[data-testid="stAppViewContainer"]::before {'
        'content:"";position:fixed;top:0;left:0;right:0;bottom:0;'
        'background-image:url("data:image/png;base64,' + LOGO_B64 + '");'
        'background-repeat:no-repeat;background-position:center center;'
        'background-size:min(95vh,95vw) auto;opacity:0.08;'
        'pointer-events:none;z-index:0;}'
        '.main .block-container{position:relative;z-index:1;}'
        '[data-testid="stSidebar"]{z-index:2;background-color:rgba(13,33,55,0.97)!important;}'
        '</style>',
        unsafe_allow_html=True,
    )

# ─────────────────────────────────────────────
# Sidebar
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div class="sb-hero">
        <div class="sb-hero-logo">ت</div>
        <h1>منصة تقييم الأداء الذكي</h1>
        <p>جامعة ديالى - قسم ضمان الجودة</p>
    </div>
    """, unsafe_allow_html=True)

    # Global search
    st.markdown('<div class="sb-label"><span class="sb-label-bar"></span>بحث سريع</div>',
                unsafe_allow_html=True)
    global_search = st.text_input("🔍 بحث عن منتسب...", key="global_search", label_visibility="collapsed",
                                   placeholder="اكتب اسم أو رقم هاتف...")

    st.markdown('<div class="sb-label"><span class="sb-label-bar"></span>القائمة الرئيسية</div>',
                unsafe_allow_html=True)
    page = st.radio("nav", [
        "📊  لوحة المعلومات",
        "👥  إدارة المنتسبين",
        "📝  التقييم الذاتي",
        "📋  تقييم المشرف",
        "📈  التقييم التراكمي",
        "🤖  التحليل الذكي",
        "📊  تقارير وإحصائيات",
        "⚙️  الإعدادات",
    ], label_visibility="collapsed")
    page = page.split("  ", 1)[-1] if "  " in page else page

    # Dark mode toggle
    st.markdown('<div class="sb-label"><span class="sb-label-bar"></span>المظهر</div>',
                unsafe_allow_html=True)
    dark_mode = st.toggle("🌙 الوضع الداكن", value=st.session_state.dark_mode, key="dm_toggle")
    if dark_mode != st.session_state.dark_mode:
        st.session_state.dark_mode = dark_mode
        st.rerun()

    st.markdown("""
    <div class="sb-credits">
        <div class="sb-credits-title">استاذ مساعد علي حسين فاضل</div>
        <div class="sb-credits-title" style="margin-top:6px;">ناهدة موسى اسماعيل</div>
    </div>
    <div class="sb-footer">
        <p>جامعة ديالى - قسم ضمان الجودة</p>
        <p>© 2026</p>
    </div>
    """, unsafe_allow_html=True)

# Load data
D = load_all()

# ═════════════════════════════════════════════
# Global Search Results (overlay)
# ═════════════════════════════════════════════
if global_search and global_search.strip():
    st.markdown('<div class="page-title">🔍 نتائج البحث</div>', unsafe_allow_html=True)
    search_results = search_associates(D["associates"], global_search, D["affiliations"], D["certificates"], D["nicknames"], D["positions"])
    if search_results:
        st.markdown(f"**تم العثور على {len(search_results)} نتيجة لـ «{global_search}»**")
        rows = []
        for a in search_results:
            rows.append({
                "#": a["id"],
                "الاسم": a["name"],
                "الهاتف": a.get("phone", ""),
                "الشهادة": lookup_name(D["certificates"], a.get("certificate_id")),
                "اللقب العلمي": lookup_name(D["nicknames"], a.get("nickname_id")),
                "المنصب": lookup_name(D["positions"], a.get("position_id")),
                "الجهة": lookup_name(D["affiliations"], a.get("affiliation_id")),
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    else:
        st.info(f"لا توجد نتائج لـ «{global_search}»")
    st.markdown("---")
    st.info("💡 أزل نص البحث من الشريط الجانبي للعودة إلى الصفحة الحالية")
    st.stop()

# ═════════════════════════════════════════════
# PAGE 1 — Dashboard
# ═════════════════════════════════════════════
if page == "لوحة المعلومات":
    st.markdown('<div class="page-title">📊 لوحة المعلومات الرئيسية</div>', unsafe_allow_html=True)

    # Welcome banner
    st.markdown(f"""
    <div class="guide-box" style="background: linear-gradient(135deg, #1e40af, #2563eb, #7c3aed); border: none;">
        <h4 style="color: #fff !important; font-size: 1.3rem !important;">مرحباً بك في منصة تقييم الأداء الذكي</h4>
        <p style="color: #dbeafe !important;">جامعة ديالى — قسم ضمان الجودة | {datetime.date.today().strftime('%Y-%m-%d')}</p>
    </div>
    """, unsafe_allow_html=True)

    associates = D["associates"]
    evaluations = D["evaluations"]
    years = D["years"]

    # Filters
    fc1, fc2 = st.columns(2)
    with fc1:
        aff_names = ["الكل"] + [a["name"] for a in D["affiliations"]]
        f_aff = st.selectbox("فلتر حسب الجهة", aff_names, key="dash_aff")
    with fc2:
        year_names = ["الكل"] + [y["year"] for y in years]
        f_year = st.selectbox("فلتر حسب السنة", year_names, key="dash_year")

    # Filter associates by affiliation
    if f_aff != "الكل":
        aff_id = next((a["id"] for a in D["affiliations"] if a["name"] == f_aff), None)
        filtered_assoc = [a for a in associates if a.get("affiliation_id") == aff_id]
    else:
        filtered_assoc = associates

    assoc_ids = {a["id"] for a in filtered_assoc}

    # Filter evaluations
    if f_year != "الكل":
        yr_id = next((y["id"] for y in years if y["year"] == f_year), None)
        filtered_eval = [e for e in evaluations if e.get("year_id") == yr_id and e.get("associate_id") in assoc_ids]
    else:
        filtered_eval = [e for e in evaluations if e.get("associate_id") in assoc_ids]

    total_emp = len(filtered_assoc)
    evaluated = len([e for e in filtered_eval if e.get("status") == "finalized"])
    pending = total_emp - evaluated if f_year != "الكل" else len([e for e in filtered_eval if e.get("status") != "finalized"])
    scores = [e["supervisor_score"] for e in filtered_eval if e.get("supervisor_score") is not None]
    avg_score = round(sum(scores) / len(scores), 1) if scores else 0

    # Completion percentage
    comp_pct = round(evaluated / total_emp * 100) if total_emp > 0 else 0

    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        st.markdown(f'<div class="stat-card" style="border-top-color:#2563eb;">'
                    f'<p style="font-size:1.8rem !important; margin-bottom:0.2rem;">👥</p>'
                    f'<h2>{total_emp}</h2><p>إجمالي المنتسبين</p></div>',
                    unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="stat-card" style="border-top-color:#16a34a;">'
                    f'<p style="font-size:1.8rem !important; margin-bottom:0.2rem;">✅</p>'
                    f'<h2 style="color:#16a34a;">{evaluated}</h2>'
                    f'<p>تم تقييمهم</p></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="stat-card" style="border-top-color:#d97706;">'
                    f'<p style="font-size:1.8rem !important; margin-bottom:0.2rem;">⏳</p>'
                    f'<h2 style="color:#d97706;">{pending}</h2>'
                    f'<p>بانتظار التقييم</p></div>', unsafe_allow_html=True)
    with c4:
        avg_grade, avg_color = score_to_grade(int(avg_score)) if avg_score else ("—", "#888")
        st.markdown(f'<div class="stat-card" style="border-top-color:{avg_color};">'
                    f'<p style="font-size:1.8rem !important; margin-bottom:0.2rem;">📈</p>'
                    f'<h2 style="color:{avg_color};">{avg_score}</h2>'
                    f'<p>المعدل ({avg_grade})</p></div>', unsafe_allow_html=True)
    with c5:
        pct_color = "#16a34a" if comp_pct >= 80 else "#d97706" if comp_pct >= 50 else "#dc2626"
        st.markdown(f'<div class="stat-card" style="border-top-color:{pct_color};">'
                    f'<p style="font-size:1.8rem !important; margin-bottom:0.2rem;">📊</p>'
                    f'<h2 style="color:{pct_color};">{comp_pct}%</h2>'
                    f'<p>نسبة الإنجاز</p></div>', unsafe_allow_html=True)

    # Progress bars
    st.markdown("---")
    pb1, pb2 = st.columns(2)
    with pb1:
        st.markdown(render_progress_bar(evaluated, total_emp, "نسبة إكمال التقييم", "#2563eb"), unsafe_allow_html=True)
    with pb2:
        excellent = sum(1 for s in scores if s >= 90) if scores else 0
        st.markdown(render_progress_bar(excellent, len(scores) if scores else 1, "نسبة المتميزين (90+)", "#16a34a"), unsafe_allow_html=True)

    st.markdown("---")

    if scores:
        # Score distribution chart
        ch1, ch2 = st.columns(2)
        with ch1:
            st.markdown("#### توزيع الدرجات")
            grade_counts = {"ممتاز": 0, "جيد جداً": 0, "جيد": 0, "مقبول": 0, "ضعيف": 0}
            for s in scores:
                g, _ = score_to_grade(s)
                if g in grade_counts:
                    grade_counts[g] += 1
            fig = px.bar(
                x=list(grade_counts.keys()),
                y=list(grade_counts.values()),
                color=list(grade_counts.keys()),
                color_discrete_map={"ممتاز": "#16a34a", "جيد جداً": "#2563eb", "جيد": "#7c3aed", "مقبول": "#d97706", "ضعيف": "#dc2626"},
                labels={"x": "التقدير", "y": "العدد"},
            )
            fig.update_layout(showlegend=False, height=350, font=dict(family="Tajawal", size=14),
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                margin=dict(l=20, r=20, t=30, b=30))
            st.plotly_chart(fig, use_container_width=True)

        with ch2:
            # Year trend
            st.markdown("#### اتجاه المعدل عبر السنوات")
            year_avgs = []
            for y in years:
                y_evals = [e["supervisor_score"] for e in evaluations
                           if e.get("year_id") == y["id"] and e.get("supervisor_score") is not None
                           and e.get("associate_id") in assoc_ids]
                if y_evals:
                    year_avgs.append({"السنة": y["year"], "المعدل": round(sum(y_evals) / len(y_evals), 1)})
            if year_avgs:
                fig2 = px.line(pd.DataFrame(year_avgs), x="السنة", y="المعدل", markers=True)
                fig2.update_layout(height=350, font=dict(family="Tajawal"))
                fig2.update_traces(line_color="#2563eb", line_width=3)
                st.plotly_chart(fig2, use_container_width=True)

        # Top / Bottom performers
        st.markdown("---")
        t1, t2 = st.columns(2)
        with t1:
            st.markdown("#### 🏆 أعلى 5 درجات")
            top5 = sorted(filtered_eval, key=lambda e: e.get("supervisor_score") or 0, reverse=True)[:5]
            for e in top5:
                name = lookup_name(associates, e["associate_id"])
                g, gc = score_to_grade(e.get("supervisor_score"))
                st.markdown(f'<div class="result-card"><h4>{name}</h4>'
                            f'<p>الدرجة: <span class="eval-badge" style="background:{gc}22;color:{gc};">{e.get("supervisor_score")} - {g}</span></p></div>',
                            unsafe_allow_html=True)
        with t2:
            st.markdown("#### ⚠️ أدنى 5 درجات")
            bot5 = sorted(filtered_eval, key=lambda e: e.get("supervisor_score") or 999)[:5]
            for e in bot5:
                name = lookup_name(associates, e["associate_id"])
                g, gc = score_to_grade(e.get("supervisor_score"))
                st.markdown(f'<div class="result-card"><h4>{name}</h4>'
                            f'<p>الدرجة: <span class="eval-badge" style="background:{gc}22;color:{gc};">{e.get("supervisor_score")} - {g}</span></p></div>',
                            unsafe_allow_html=True)
    else:
        st.info("لا توجد بيانات تقييم لعرضها. ابدأ بإضافة منتسبين وتقييمهم.")


# ═════════════════════════════════════════════
# PAGE 2 — Staff Management
# ═════════════════════════════════════════════
elif page == "إدارة المنتسبين":
    st.markdown('<div class="page-title">إدارة المنتسبين</div>', unsafe_allow_html=True)

    associates = D["associates"]

    tab1, tab2, tab3 = st.tabs(["📋 قائمة المنتسبين", "➕ إضافة منتسب", "📤 استيراد من Excel"])

    with tab1:
        # Filters
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            f_aff = st.selectbox("الجهة", ["الكل"] + [a["name"] for a in D["affiliations"]], key="staff_aff")
        with fc2:
            f_pos = st.selectbox("المنصب", ["الكل"] + [p["name"] for p in D["positions"]], key="staff_pos")
        with fc3:
            f_search = st.text_input("بحث بالاسم", key="staff_search")

        filtered = associates.copy()
        if f_aff != "الكل":
            aid = next((a["id"] for a in D["affiliations"] if a["name"] == f_aff), None)
            filtered = [a for a in filtered if a.get("affiliation_id") == aid]
        if f_pos != "الكل":
            pid = next((p["id"] for p in D["positions"] if p["name"] == f_pos), None)
            filtered = [a for a in filtered if a.get("position_id") == pid]
        if f_search:
            filtered = [a for a in filtered if f_search in a.get("name", "")]

        if filtered:
            st.markdown(f"**إجمالي: {len(filtered)} منتسب**")

            rows = []
            for a in filtered:
                rows.append({
                    "#": a["id"],
                    "الاسم": a["name"],
                    "الهاتف": a.get("phone", ""),
                    "الشهادة": lookup_name(D["certificates"], a.get("certificate_id")),
                    "اللقب العلمي": lookup_name(D["nicknames"], a.get("nickname_id")),
                    "المنصب": lookup_name(D["positions"], a.get("position_id")),
                    "الجهة": lookup_name(D["affiliations"], a.get("affiliation_id")),
                })
            df = pd.DataFrame(rows)
            st.dataframe(df, use_container_width=True, hide_index=True)

            # Export
            excel_data = export_df_excel(df, "المنتسبين")
            if excel_data:
                st.download_button("📥 تصدير إلى Excel", data=excel_data,
                                   file_name=f"المنتسبين_{datetime.date.today()}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            # Delete
            st.markdown("---")
            st.markdown("#### حذف منتسب")
            del_names = {a["name"]: a["id"] for a in associates}
            del_sel = st.selectbox("اختر المنتسب للحذف", [""] + list(del_names.keys()), key="del_assoc")
            if del_sel and st.button("🗑️ حذف", type="secondary", key="btn_del_assoc"):
                del_id = del_names[del_sel]
                new_assoc = [a for a in associates if a["id"] != del_id]
                save_json("associates.json", new_assoc)
                # Also remove their evaluations
                new_evals = [e for e in D["evaluations"] if e.get("associate_id") != del_id]
                save_json("evaluations.json", new_evals)
                st.success(f"تم حذف {del_sel}")
                st.rerun()
        else:
            st.info("لا توجد نتائج")

    with tab2:
        st.markdown("""
        <div class="guide-box">
            <h4>إضافة منتسب جديد</h4>
            <p>أدخل بيانات المنتسب في الحقول أدناه ثم اضغط "حفظ"</p>
        </div>
        """, unsafe_allow_html=True)

        with st.form("add_associate", clear_on_submit=True):
            a_name = st.text_input("الاسم الكامل *")
            a_phone = st.text_input("رقم الهاتف *")
            ac1, ac2 = st.columns(2)
            with ac1:
                a_cert = st.selectbox("الشهادة", [c["name"] for c in D["certificates"]])
                a_nick = st.selectbox("اللقب العلمي", [n["name"] for n in D["nicknames"]])
            with ac2:
                a_pos = st.selectbox("المنصب", [p["name"] for p in D["positions"]])
                a_aff = st.selectbox("الجهة", [a["name"] for a in D["affiliations"]])

            submitted = st.form_submit_button("💾 حفظ", type="primary", use_container_width=True)
            if submitted:
                if not a_name or not a_phone:
                    st.error("الاسم ورقم الهاتف مطلوبان")
                else:
                    new = {
                        "id": next_id(associates),
                        "name": a_name.strip(),
                        "phone": a_phone.strip(),
                        "certificate_id": next((c["id"] for c in D["certificates"] if c["name"] == a_cert), 1),
                        "nickname_id": next((n["id"] for n in D["nicknames"] if n["name"] == a_nick), 1),
                        "position_id": next((p["id"] for p in D["positions"] if p["name"] == a_pos), 1),
                        "affiliation_id": next((a["id"] for a in D["affiliations"] if a["name"] == a_aff), 1),
                    }
                    associates.append(new)
                    save_json("associates.json", associates)
                    st.success(f"تم إضافة {a_name}")
                    st.rerun()

    with tab3:
        st.markdown("""
        <div class="guide-box">
            <h4>استيراد من Excel</h4>
            <p>يجب أن يحتوي الملف على الأعمدة: الاسم، الهاتف، الشهادة، اللقب العلمي، المنصب، الجهة</p>
        </div>
        """, unsafe_allow_html=True)

        uploaded = st.file_uploader("ارفع ملف Excel", type=["xlsx", "xls"], key="import_staff")
        if uploaded:
            try:
                imp_df = pd.read_excel(uploaded, sheet_name=0)
                st.dataframe(imp_df, use_container_width=True, hide_index=True)

                if st.button("📥 استيراد البيانات", type="primary"):
                    count = 0
                    for _, row in imp_df.iterrows():
                        name = str(row.get("الاسم", row.get("affiliatename", ""))).strip()
                        if not name or name.lower() in ("nan", "none"):
                            continue
                        phone = str(row.get("الهاتف", row.get("Phonenumber", ""))).strip()
                        cert_name = str(row.get("الشهادة", "بكالوريوس")).strip()
                        nick_name = str(row.get("اللقب العلمي", "بلا لقب")).strip()
                        pos_name = str(row.get("المنصب", "موظف")).strip()
                        aff_name = str(row.get("الجهة", "")).strip()

                        new = {
                            "id": next_id(associates),
                            "name": name,
                            "phone": phone if phone.lower() not in ("nan", "none") else "",
                            "certificate_id": next((c["id"] for c in D["certificates"] if c["name"] == cert_name), 4),
                            "nickname_id": next((n["id"] for n in D["nicknames"] if n["name"] == nick_name), 5),
                            "position_id": next((p["id"] for p in D["positions"] if p["name"] == pos_name), 8),
                            "affiliation_id": next((a["id"] for a in D["affiliations"] if a["name"] == aff_name), 1),
                        }
                        associates.append(new)
                        count += 1
                    save_json("associates.json", associates)
                    st.success(f"تم استيراد {count} منتسب بنجاح")
                    st.rerun()
            except Exception as e:
                st.error(f"خطأ في قراءة الملف: {e}")


# ═════════════════════════════════════════════
# PAGE 3 — Self-Evaluation
# ═════════════════════════════════════════════
elif page == "التقييم الذاتي":
    st.markdown('<div class="page-title">التقييم الذاتي للمنتسب</div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="guide-box">
        <h4>آلية التقييم الذاتي</h4>
        <p>① اختر المنتسب والسنة التقييمية</p>
        <p>② يدخل المنتسب درجته الذاتية (0-100) مع ملاحظات اختيارية</p>
        <p>③ بعد الحفظ، ينتقل التقييم إلى المشرف للمراجعة والتعديل</p>
    </div>
    """, unsafe_allow_html=True)

    associates = D["associates"]
    years = D["years"]
    evaluations = D["evaluations"]

    if not associates:
        st.warning("لا يوجد منتسبون. أضف منتسبين أولاً من صفحة إدارة المنتسبين.")
    elif not years:
        st.warning("لا توجد سنوات تقييمية. أضف سنة من الإعدادات.")
    else:
        sc1, sc2 = st.columns(2)
        with sc1:
            sel_assoc = st.selectbox("اختر المنتسب", [f"{a['name']} (#{a['id']})" for a in associates], key="self_assoc")
        with sc2:
            sel_year = st.selectbox("السنة التقييمية", [y["year"] for y in years], key="self_year")

        if sel_assoc and sel_year:
            assoc_id = int(sel_assoc.split("#")[1].rstrip(")"))
            year_id = next((y["id"] for y in years if y["year"] == sel_year), None)

            # Find existing evaluation
            existing = next((e for e in evaluations
                             if e.get("associate_id") == assoc_id and e.get("year_id") == year_id), None)

            if existing and existing.get("self_score") is not None:
                g, gc = score_to_grade(existing["self_score"])
                status_text = "تم التقييم الذاتي ✓" if existing.get("status") == "self_done" else "مكتمل ✓✓" if existing.get("status") == "finalized" else "قيد المراجعة"
                st.markdown(f"""
                <div class="result-card">
                    <h4>التقييم الذاتي - {sel_year}</h4>
                    <p>الدرجة الذاتية: <span class="eval-badge" style="background:{gc}22;color:{gc};">{existing['self_score']} - {g}</span></p>
                    <p>الملاحظات: {existing.get('self_notes', '—') or '—'}</p>
                    <p>الحالة: {status_text}</p>
                </div>
                """, unsafe_allow_html=True)

                if existing.get("supervisor_score") is not None:
                    sg, sgc = score_to_grade(existing["supervisor_score"])
                    st.markdown(f"""
                    <div class="result-card" style="border-right-color: {sgc};">
                        <h4>تقييم المشرف</h4>
                        <p>درجة المشرف: <span class="eval-badge" style="background:{sgc}22;color:{sgc};">{existing['supervisor_score']} - {sg}</span></p>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                with st.form("self_eval_form"):
                    self_score = st.slider("الدرجة الذاتية", 0, 100, 75, key="self_score_slider")
                    self_notes = st.text_area("ملاحظات (اختياري)", key="self_notes")

                    g, gc = score_to_grade(self_score)
                    st.markdown(f'<p>التقدير: <span class="eval-badge" style="background:{gc}22;color:{gc};">{g}</span></p>',
                                unsafe_allow_html=True)

                    if st.form_submit_button("💾 حفظ التقييم الذاتي", type="primary", use_container_width=True):
                        if existing:
                            existing["self_score"] = self_score
                            existing["self_notes"] = self_notes
                            existing["status"] = "self_done"
                        else:
                            evaluations.append({
                                "id": next_id(evaluations),
                                "associate_id": assoc_id,
                                "year_id": year_id,
                                "self_score": self_score,
                                "supervisor_score": None,
                                "self_notes": self_notes,
                                "supervisor_notes": "",
                                "status": "self_done",
                            })
                        save_json("evaluations.json", evaluations)
                        st.success("تم حفظ التقييم الذاتي بنجاح")
                        st.rerun()


# ═════════════════════════════════════════════
# PAGE 4 — Supervisor Evaluation
# ═════════════════════════════════════════════
elif page == "تقييم المشرف":
    st.markdown('<div class="page-title">تقييم المشرف</div>', unsafe_allow_html=True)

    associates = D["associates"]
    years = D["years"]
    evaluations = D["evaluations"]

    if not years:
        st.warning("لا توجد سنوات تقييمية.")
    else:
        sel_year = st.selectbox("السنة التقييمية", [y["year"] for y in years], key="sup_year")
        year_id = next((y["id"] for y in years if y["year"] == sel_year), None)

        # Filter by affiliation
        f_aff = st.selectbox("فلتر الجهة", ["الكل"] + [a["name"] for a in D["affiliations"]], key="sup_aff")
        if f_aff != "الكل":
            aff_id = next((a["id"] for a in D["affiliations"] if a["name"] == f_aff), None)
            display_assoc = [a for a in associates if a.get("affiliation_id") == aff_id]
        else:
            display_assoc = associates

        if not display_assoc:
            st.info("لا يوجد منتسبون لعرضهم")
        else:
            st.markdown("---")

            for assoc in display_assoc:
                existing = next((e for e in evaluations
                                 if e.get("associate_id") == assoc["id"] and e.get("year_id") == year_id), None)

                self_score = existing.get("self_score") if existing else None
                sup_score = existing.get("supervisor_score") if existing else None
                status = existing.get("status", "pending") if existing else "pending"

                # Color coding
                if sup_score is not None:
                    _, border_color = score_to_grade(sup_score)
                elif self_score is not None:
                    border_color = "#d97706"
                else:
                    border_color = "#888"

                self_display = f"{self_score}" if self_score is not None else "لم يُقيّم ذاتياً"
                sup_display = f"{sup_score}" if sup_score is not None else "—"

                aff_name = lookup_name(D["affiliations"], assoc.get("affiliation_id"))
                pos_name = lookup_name(D["positions"], assoc.get("position_id"))

                col1, col2 = st.columns([3, 1])
                with col1:
                    st.markdown(f"""
                    <div class="result-card" style="border-right-color: {border_color};">
                        <h4>{assoc['name']}</h4>
                        <p>{aff_name} | {pos_name}</p>
                        <p>التقييم الذاتي: <strong>{self_display}</strong> | تقييم المشرف: <strong>{sup_display}</strong></p>
                    </div>
                    """, unsafe_allow_html=True)

                with col2:
                    new_score = st.number_input(
                        "الدرجة", min_value=0, max_value=100,
                        value=sup_score if sup_score is not None else (self_score if self_score is not None else 75),
                        key=f"sup_{assoc['id']}_{year_id}"
                    )
                    if st.button("💾", key=f"save_sup_{assoc['id']}_{year_id}"):
                        if existing:
                            existing["supervisor_score"] = new_score
                            existing["status"] = "finalized"
                        else:
                            evaluations.append({
                                "id": next_id(evaluations),
                                "associate_id": assoc["id"],
                                "year_id": year_id,
                                "self_score": None,
                                "supervisor_score": new_score,
                                "self_notes": "",
                                "supervisor_notes": "",
                                "status": "finalized",
                            })
                        save_json("evaluations.json", evaluations)
                        st.success(f"تم حفظ تقييم {assoc['name']}")
                        st.rerun()


# ═════════════════════════════════════════════
# PAGE 5 — Cumulative Report
# ═════════════════════════════════════════════
elif page == "التقييم التراكمي":
    st.markdown('<div class="page-title">التقييم التراكمي</div>', unsafe_allow_html=True)

    associates = D["associates"]
    years = D["years"]
    evaluations = D["evaluations"]

    if not associates:
        st.warning("لا يوجد منتسبون")
    else:
        sel_assoc = st.selectbox("اختر المنتسب", [f"{a['name']} (#{a['id']})" for a in associates], key="cum_assoc")
        assoc_id = int(sel_assoc.split("#")[1].rstrip(")"))
        assoc = next((a for a in associates if a["id"] == assoc_id), None)

        # Get all evaluations for this person
        person_evals = [e for e in evaluations if e.get("associate_id") == assoc_id and e.get("supervisor_score") is not None]

        if assoc:
            st.markdown(f"""
            <div class="result-card">
                <h4>{assoc['name']}</h4>
                <p>الجهة: {lookup_name(D['affiliations'], assoc.get('affiliation_id'))} | المنصب: {lookup_name(D['positions'], assoc.get('position_id'))}</p>
                <p>الشهادة: {lookup_name(D['certificates'], assoc.get('certificate_id'))} | اللقب: {lookup_name(D['nicknames'], assoc.get('nickname_id'))}</p>
            </div>
            """, unsafe_allow_html=True)

        if person_evals:
            # Table
            rows = []
            for e in person_evals:
                yr_name = lookup_name(years, e["year_id"], "year")
                g, gc = score_to_grade(e["supervisor_score"])
                sg = score_to_grade(e.get("self_score"))[0] if e.get("self_score") is not None else "—"
                rows.append({
                    "السنة": yr_name,
                    "التقييم الذاتي": e.get("self_score", "—"),
                    "تقييم المشرف": e["supervisor_score"],
                    "التقدير": g,
                })

            cum_avg = round(sum(e["supervisor_score"] for e in person_evals) / len(person_evals), 1)
            cum_grade, cum_color = score_to_grade(int(cum_avg))

            c1, c2, c3 = st.columns(3)
            with c1:
                st.markdown(f'<div class="stat-card"><h2>{len(person_evals)}</h2><p>سنوات التقييم</p></div>',
                            unsafe_allow_html=True)
            with c2:
                st.markdown(f'<div class="stat-card"><h2 style="color:{cum_color};">{cum_avg}</h2>'
                            f'<p>المعدل التراكمي</p></div>', unsafe_allow_html=True)
            with c3:
                st.markdown(f'<div class="stat-card"><h2 style="color:{cum_color};">{cum_grade}</h2>'
                            f'<p>التقدير التراكمي</p></div>', unsafe_allow_html=True)

            st.markdown("---")
            df = pd.DataFrame(rows)
            st.dataframe(df, use_container_width=True, hide_index=True)

            # Chart: self vs supervisor over years
            chart_data = []
            for e in person_evals:
                yr_name = lookup_name(years, e["year_id"], "year")
                chart_data.append({"السنة": yr_name, "النوع": "تقييم المشرف", "الدرجة": e["supervisor_score"]})
                if e.get("self_score") is not None:
                    chart_data.append({"السنة": yr_name, "النوع": "التقييم الذاتي", "الدرجة": e["self_score"]})

            if chart_data:
                fig = px.line(pd.DataFrame(chart_data), x="السنة", y="الدرجة", color="النوع", markers=True,
                              color_discrete_map={"تقييم المشرف": "#2563eb", "التقييم الذاتي": "#d97706"})
                fig.update_layout(height=400, font=dict(family="Tajawal", size=14),
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                margin=dict(l=20, r=20, t=30, b=30))
                st.plotly_chart(fig, use_container_width=True)

            # Export
            exp1, exp2 = st.columns(2)
            with exp1:
                excel_data = export_df_excel(df, f"تقييم {assoc['name']}")
                if excel_data:
                    st.download_button("📥 تصدير Excel", data=excel_data,
                                       file_name=f"تقييم_تراكمي_{assoc['name']}_{datetime.date.today()}.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       type="primary", use_container_width=True)
            with exp2:
                pdf_data = generate_employee_pdf(assoc, D["evaluations"], D["years"],
                                                  D["affiliations"], D["certificates"], D["nicknames"], D["positions"])
                if pdf_data:
                    st.download_button("📄 تصدير PDF", data=pdf_data,
                                       file_name=f"تقرير_{assoc['name']}_{datetime.date.today()}.pdf",
                                       mime="application/pdf",
                                       use_container_width=True)
        else:
            st.info("لا توجد تقييمات لهذا المنتسب بعد")


# ═════════════════════════════════════════════
# PAGE 6 — AI Smart Analysis
# ═════════════════════════════════════════════
elif page == "التحليل الذكي":
    st.markdown('<div class="page-title">🤖 التحليل الذكي بالذكاء الاصطناعي</div>', unsafe_allow_html=True)

    associates = D["associates"]
    years = D["years"]
    evaluations = D["evaluations"]

    if not associates or not evaluations:
        st.warning("لا توجد بيانات كافية للتحليل. أضف منتسبين وقيّمهم أولاً.")
    else:
        # ── Section 1: Smart Ranking ──
        st.markdown("### 🏆 التصنيف الذكي للمنتسبين")
        st.markdown("""
        <div class="guide-box">
            <h4>كيف يعمل التصنيف الذكي؟</h4>
            <p>يعتمد على ثلاثة عوامل: <strong>المعدل (50%)</strong> + <strong>الثبات (25%)</strong> + <strong>معدل التحسن (25%)</strong></p>
            <p>هذا أفضل من الترتيب بالدرجة فقط — يكافئ من يتحسن باستمرار</p>
        </div>
        """, unsafe_allow_html=True)

        rankings = ai_smart_rank(associates, evaluations, years)
        if rankings:
            for rank_idx, r in enumerate(rankings, 1):
                medal = "🥇" if rank_idx == 1 else "🥈" if rank_idx == 2 else "🥉" if rank_idx == 3 else f"#{rank_idx}"
                g, gc = score_to_grade(int(r["avg"]))
                st.markdown(f"""
                <div class="result-card" style="border-right-color: {r['cat_color']};">
                    <h4>{medal} {r['name']}</h4>
                    <p>
                        المعدل: <span class="eval-badge" style="background:{gc}22;color:{gc};">{r['avg']} - {g}</span>
                        &nbsp;&nbsp;|&nbsp;&nbsp;
                        الثبات: <strong>{r['consistency']}</strong>
                        &nbsp;&nbsp;|&nbsp;&nbsp;
                        التحسن: <strong>{r['improvement']:+.1f}</strong>
                        &nbsp;&nbsp;|&nbsp;&nbsp;
                        النتيجة المركبة: <strong>{r['composite']}</strong>
                    </p>
                    <p>التصنيف: <span class="eval-badge" style="background:{r['cat_color']}22;color:{r['cat_color']};">{r['category']}</span></p>
                </div>
                """, unsafe_allow_html=True)

        st.markdown("---")

        # ── Section 2: Individual AI Analysis ──
        st.markdown("### 📊 التحليل الفردي الذكي")
        sel_assoc = st.selectbox("اختر المنتسب", [f"{a['name']} (#{a['id']})" for a in associates], key="ai_assoc")
        assoc_id = int(sel_assoc.split("#")[1].rstrip(")"))
        assoc = next((a for a in associates if a["id"] == assoc_id), None)

        person_evals = sorted(
            [e for e in evaluations if e.get("associate_id") == assoc_id and e.get("supervisor_score") is not None],
            key=lambda e: e.get("year_id", 0)
        )

        if not person_evals:
            st.info("لا توجد تقييمات لهذا المنتسب بعد")
        else:
            scores_by_year = []
            for e in person_evals:
                yr_name = lookup_name(years, e["year_id"], "year")
                scores_by_year.append((yr_name, e["supervisor_score"]))

            avg_score = sum(s for _, s in scores_by_year) / len(scores_by_year)

            # Trend Analysis
            trend = ai_trend_analysis(scores_by_year)

            # Prediction
            predicted = ai_predict_next(scores_by_year)

            # Self-evaluation gap (use latest)
            latest_eval = person_evals[-1]
            gap_alert = ai_self_gap_alert(latest_eval.get("self_score"), latest_eval.get("supervisor_score"))

            # Find ranking category
            rank_entry = next((r for r in rankings if r["id"] == assoc_id), None) if rankings else None
            category = rank_entry["category"] if rank_entry else ""

            # Recommendations
            recs = ai_recommendations(assoc["name"], avg_score, trend, gap_alert, category)

            # Display
            col1, col2, col3 = st.columns(3)
            with col1:
                st.markdown(f"""
                <div class="stat-card">
                    <h2 style="font-size:2.5rem;">{trend['icon']}</h2>
                    <h2 style="color:{trend['color']}; font-size:1.3rem;">{trend['trend']}</h2>
                    <p>{trend['description']}</p>
                </div>
                """, unsafe_allow_html=True)
            with col2:
                if predicted is not None:
                    pg, pgc = score_to_grade(predicted)
                    st.markdown(f"""
                    <div class="stat-card">
                        <h2 style="font-size:2.5rem;">🔮</h2>
                        <h2 style="color:{pgc};">{predicted}</h2>
                        <p>التوقع للسنة القادمة ({pg})</p>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown("""
                    <div class="stat-card">
                        <h2 style="font-size:2.5rem;">🔮</h2>
                        <h2 style="color:#888;">—</h2>
                        <p>بيانات غير كافية للتنبؤ</p>
                    </div>
                    """, unsafe_allow_html=True)
            with col3:
                if gap_alert:
                    st.markdown(f"""
                    <div class="stat-card">
                        <h2 style="font-size:2.5rem;">{gap_alert['icon']}</h2>
                        <h2 style="color:{gap_alert['color']}; font-size:1rem;">{gap_alert['msg']}</h2>
                        <p>فجوة التقييم الذاتي</p>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown("""
                    <div class="stat-card">
                        <h2 style="font-size:2.5rem;">📝</h2>
                        <h2 style="color:#888;">—</h2>
                        <p>لا يوجد تقييم ذاتي للمقارنة</p>
                    </div>
                    """, unsafe_allow_html=True)

            # Prediction chart
            if predicted is not None and len(scores_by_year) >= 2:
                st.markdown("---")
                st.markdown("#### 📈 مسار الأداء والتوقع")
                chart_rows = []
                for yr, sc in scores_by_year:
                    chart_rows.append({"السنة": yr, "الدرجة": sc, "النوع": "فعلي"})
                chart_rows.append({"السنة": "التوقع القادم", "الدرجة": predicted, "النوع": "متوقع"})
                fig = px.line(pd.DataFrame(chart_rows), x="السنة", y="الدرجة", color="النوع",
                              markers=True,
                              color_discrete_map={"فعلي": "#2563eb", "متوقع": "#d97706"})
                fig.update_layout(height=350, font=dict(family="Tajawal", size=14),
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                margin=dict(l=20, r=20, t=30, b=30))
                st.plotly_chart(fig, use_container_width=True)

            # Recommendations
            st.markdown("---")
            st.markdown("#### 💡 التوصيات الذكية")
            for rec in recs:
                st.markdown(f"""
                <div class="result-card" style="border-right-color: #2563eb; padding: 0.8rem;">
                    <p style="margin:0; font-size: 1.1rem;">{rec}</p>
                </div>
                """, unsafe_allow_html=True)

        # ── Section 3: Alerts Summary ──
        st.markdown("---")
        st.markdown("### ⚠️ التنبيهات الذكية")

        alerts = []
        for a in associates:
            p_evals = sorted(
                [e for e in evaluations if e.get("associate_id") == a["id"] and e.get("supervisor_score") is not None],
                key=lambda e: e.get("year_id", 0)
            )
            if len(p_evals) >= 2:
                scores_list = [(lookup_name(years, e["year_id"], "year"), e["supervisor_score"]) for e in p_evals]
                t = ai_trend_analysis(scores_list)
                if t["trend"] in ("تراجع طفيف", "تراجع حاد"):
                    alerts.append({"name": a["name"], "alert": t["description"], "icon": t["icon"], "color": t["color"]})

            for e in p_evals:
                gap = ai_self_gap_alert(e.get("self_score"), e.get("supervisor_score"))
                if gap and gap["type"] in ("overconfidence",):
                    alerts.append({"name": a["name"], "alert": gap["msg"], "icon": gap["icon"], "color": gap["color"]})
                    break

        if alerts:
            for al in alerts:
                st.markdown(f"""
                <div class="result-card" style="border-right-color: {al['color']};">
                    <h4>{al['icon']} {al['name']}</h4>
                    <p>{al['alert']}</p>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.success("✅ لا توجد تنبيهات — أداء جميع المنتسبين مستقر أو متحسن")


# ═════════════════════════════════════════════
# PAGE 7 — Reports & Statistics
# ═════════════════════════════════════════════
elif page == "تقارير وإحصائيات":
    st.markdown('<div class="page-title">تقارير وإحصائيات</div>', unsafe_allow_html=True)

    associates = D["associates"]
    years = D["years"]
    evaluations = D["evaluations"]

    sel_year = st.selectbox("السنة", [y["year"] for y in years] if years else ["لا توجد سنوات"], key="rep_year")
    year_id = next((y["id"] for y in years if y["year"] == sel_year), None)

    if year_id:
        year_evals = [e for e in evaluations if e.get("year_id") == year_id and e.get("supervisor_score") is not None]

        if year_evals:
            # Department comparison
            st.markdown("#### مقارنة الجهات")
            dept_data = {}
            for e in year_evals:
                assoc = next((a for a in associates if a["id"] == e["associate_id"]), None)
                if assoc:
                    aff = lookup_name(D["affiliations"], assoc.get("affiliation_id"))
                    dept_data.setdefault(aff, []).append(e["supervisor_score"])

            dept_rows = [{"الجهة": k, "المعدل": round(sum(v)/len(v), 1), "العدد": len(v)}
                         for k, v in dept_data.items()]
            dept_rows.sort(key=lambda x: x["المعدل"], reverse=True)

            if dept_rows:
                fig = px.bar(pd.DataFrame(dept_rows), x="الجهة", y="المعدل", color="المعدل",
                             color_continuous_scale=["#dc2626", "#d97706", "#16a34a"],
                             text="المعدل")
                fig.update_layout(height=400, font=dict(family="Tajawal", size=14), showlegend=False,
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                margin=dict(l=20, r=20, t=30, b=30))
                fig.update_traces(textposition="outside")
                st.plotly_chart(fig, use_container_width=True)

            st.markdown("---")

            # Score histogram
            st.markdown("#### توزيع الدرجات")
            scores = [e["supervisor_score"] for e in year_evals]
            fig2 = px.histogram(x=scores, nbins=10, labels={"x": "الدرجة", "y": "العدد"},
                                color_discrete_sequence=["#2563eb"])
            fig2.update_layout(height=350, font=dict(family="Tajawal"))
            st.plotly_chart(fig2, use_container_width=True)

            st.markdown("---")

            # Comprehensive table
            st.markdown("#### الجدول الشامل")
            comp_rows = []
            for e in year_evals:
                assoc = next((a for a in associates if a["id"] == e["associate_id"]), None)
                if assoc:
                    g, _ = score_to_grade(e["supervisor_score"])
                    sg, _ = score_to_grade(e.get("self_score"))
                    comp_rows.append({
                        "الاسم": assoc["name"],
                        "الجهة": lookup_name(D["affiliations"], assoc.get("affiliation_id")),
                        "المنصب": lookup_name(D["positions"], assoc.get("position_id")),
                        "التقييم الذاتي": e.get("self_score", "—"),
                        "تقييم المشرف": e["supervisor_score"],
                        "التقدير": g,
                        "الفرق": (e["supervisor_score"] - e["self_score"]) if e.get("self_score") is not None else "—",
                    })

            comp_df = pd.DataFrame(comp_rows)
            st.dataframe(comp_df, use_container_width=True, hide_index=True)

            # Export
            excel_data = export_df_excel(comp_df, f"تقرير {sel_year}")
            if excel_data:
                st.download_button("📥 تصدير التقرير الشامل", data=excel_data,
                                   file_name=f"تقرير_شامل_{sel_year}_{datetime.date.today()}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   type="primary", use_container_width=True)
        else:
            st.info("لا توجد تقييمات لهذه السنة")

        # Year-over-year trend for all departments
        st.markdown("---")
        st.markdown("#### اتجاه الأداء عبر السنوات (جميع الجهات)")
        trend_data = []
        for y in years:
            y_evals = [e for e in evaluations if e.get("year_id") == y["id"] and e.get("supervisor_score") is not None]
            dept_map = {}
            for e in y_evals:
                assoc = next((a for a in associates if a["id"] == e["associate_id"]), None)
                if assoc:
                    aff = lookup_name(D["affiliations"], assoc.get("affiliation_id"))
                    dept_map.setdefault(aff, []).append(e["supervisor_score"])
            for dept, scores in dept_map.items():
                trend_data.append({"السنة": y["year"], "الجهة": dept, "المعدل": round(sum(scores)/len(scores), 1)})

        if trend_data:
            fig3 = px.line(pd.DataFrame(trend_data), x="السنة", y="المعدل", color="الجهة", markers=True)
            fig3.update_layout(height=400, font=dict(family="Tajawal"))
            st.plotly_chart(fig3, use_container_width=True)


# ═════════════════════════════════════════════
# PAGE 7 — Settings
# ═════════════════════════════════════════════
elif page == "الإعدادات":
    st.markdown('<div class="page-title">⚙️ الإعدادات — لوحة المشرف</div>', unsafe_allow_html=True)

    # Admin authentication
    ADMIN_PASSWORD = "admin2026"

    if "admin_authenticated" not in st.session_state:
        st.session_state.admin_authenticated = False

    if not st.session_state.admin_authenticated:
        st.markdown("""
        <div class="guide-box">
            <h4>🔐 تسجيل دخول المشرف</h4>
            <p>هذه الصفحة محمية. أدخل كلمة مرور المشرف للوصول إلى إعدادات جداول الترميز.</p>
        </div>
        """, unsafe_allow_html=True)

        with st.form("admin_login"):
            pwd = st.text_input("كلمة المرور", type="password", key="admin_pwd")
            if st.form_submit_button("🔓 دخول", type="primary", use_container_width=True):
                if pwd == ADMIN_PASSWORD:
                    st.session_state.admin_authenticated = True
                    st.rerun()
                else:
                    st.error("❌ كلمة المرور غير صحيحة")
        st.stop()

    # Admin is authenticated — show logout button
    if st.button("🔒 تسجيل خروج المشرف", key="admin_logout"):
        st.session_state.admin_authenticated = False
        st.rerun()

    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "📅 السنوات التقييمية", "🏛️ الجهات", "📜 الشهادات",
        "🎓 الألقاب العلمية", "💼 المناصب", "💾 نسخ احتياطي"
    ])

    # ── Evaluation Years ──
    with tab1:
        st.markdown("#### السنوات التقييمية الحالية")
        for y in D["years"]:
            st.markdown(f'<div class="result-card"><h4>{y["year"]}</h4></div>', unsafe_allow_html=True)

        with st.form("add_year"):
            new_year = st.text_input("سنة جديدة (مثال: 2025-2026)")
            if st.form_submit_button("➕ إضافة سنة", type="primary"):
                if new_year.strip():
                    yrs = D["years"]
                    yrs.append({"id": next_id(yrs), "year": new_year.strip()})
                    save_json("evaluation_years.json", yrs)
                    st.success(f"تم إضافة {new_year}")
                    st.rerun()

    # Helper for lookup table management
    def manage_lookup(tab, items, file_name, label, key_field="name"):
        with tab:
            st.markdown(f"#### {label} الحالية")
            for it in items:
                st.markdown(f"- {it[key_field]}")

            with st.form(f"add_{file_name}"):
                new_val = st.text_input(f"إضافة {label} جديدة")
                if st.form_submit_button(f"➕ إضافة", type="primary"):
                    if new_val.strip():
                        items.append({"id": next_id(items), key_field: new_val.strip()})
                        save_json(file_name, items)
                        st.success(f"تم إضافة {new_val}")
                        st.rerun()

    manage_lookup(tab2, D["affiliations"], "affiliations.json", "الجهات")
    manage_lookup(tab3, D["certificates"], "certificates.json", "الشهادات")
    manage_lookup(tab4, D["nicknames"], "nicknames.json", "الألقاب العلمية")
    manage_lookup(tab5, D["positions"], "positions.json", "المناصب")

    # ── Backup / Restore ──
    with tab6:
        st.markdown("#### تصدير نسخة احتياطية")
        if st.button("📦 تصدير جميع البيانات", type="primary"):
            backup = {
                "affiliations": D["affiliations"],
                "certificates": D["certificates"],
                "nicknames": D["nicknames"],
                "positions": D["positions"],
                "associates": D["associates"],
                "evaluation_years": D["years"],
                "evaluations": D["evaluations"],
                "backup_date": str(datetime.datetime.now()),
            }
            backup_json = json.dumps(backup, ensure_ascii=False, indent=2)
            st.download_button("💾 تحميل النسخة الاحتياطية",
                               data=backup_json.encode("utf-8"),
                               file_name=f"backup_{datetime.date.today()}.json",
                               mime="application/json")

        st.markdown("---")
        st.markdown("#### استعادة من نسخة احتياطية")
        restore_file = st.file_uploader("ارفع ملف النسخة الاحتياطية (JSON)", type=["json"], key="restore")
        if restore_file:
            try:
                backup = json.loads(restore_file.read().decode("utf-8"))
                st.json({"تاريخ النسخة": backup.get("backup_date", "غير معروف"),
                         "المنتسبين": len(backup.get("associates", [])),
                         "التقييمات": len(backup.get("evaluations", []))})

                if st.button("♻️ استعادة البيانات", type="primary"):
                    file_map = {
                        "affiliations": "affiliations.json",
                        "certificates": "certificates.json",
                        "nicknames": "nicknames.json",
                        "positions": "positions.json",
                        "associates": "associates.json",
                        "evaluation_years": "evaluation_years.json",
                        "evaluations": "evaluations.json",
                    }
                    for key, fname in file_map.items():
                        if key in backup:
                            save_json(fname, backup[key])
                    st.success("تم استعادة البيانات بنجاح!")
                    st.rerun()
            except Exception as e:
                st.error(f"خطأ في قراءة الملف: {e}")
